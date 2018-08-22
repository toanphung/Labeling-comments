import openpyxl

print("Loading excel file...")
wb = openpyxl.load_workbook('./foody-data-1531-26072018.xlsx')

comments_sheet_name = wb.sheetnames[1]

comments_worksheet = wb[comments_sheet_name]

print("Done loading excel file!")

print("Start processing...")

with open('comments_data_foody.txt', 'w') as f:
	for row in range(2, comments_worksheet.max_row + 1):
		for column in "I":
			cell_name = "{}{}".format(column, row)
			cell_val = comments_worksheet[cell_name].value
			if cell_val is None:
				cell_val =  ""
			f.write(str(cell_val)+"\n")

print("Process successfully!")
